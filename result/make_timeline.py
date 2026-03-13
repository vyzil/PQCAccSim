import re
import math
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series

def resolve_paths():
    parser = argparse.ArgumentParser(description="Build timeline workbook from simulator trace log")
    parser.add_argument(
        "--log",
        default=None,
        help="Input log path (default: auto by --level, else result.log)",
    )
    parser.add_argument(
        "--out",
        default=None,
        help="Output xlsx path (default: derived from log/level)",
    )
    parser.add_argument(
        "--level",
        default=None,
        choices=["default", "2", "3", "5"],
        help="Optional mode hint used for default naming",
    )
    args = parser.parse_args()

    if args.log:
        log_path = Path(args.log)
    elif args.level is not None:
        tag = f"d{args.level}" if args.level in {"2", "3", "5"} else "default"
        log_path = Path(f"result_{tag}.log")
    else:
        log_path = Path("result.log")

    if not log_path.is_absolute() and not log_path.exists():
        alt = Path("result") / log_path
        if alt.exists():
            log_path = alt

    if args.out:
        out_path = Path(args.out)
    else:
        stem = log_path.stem
        suffix = stem[len("result_"):] if stem.startswith("result_") else stem
        if suffix in {"result", ""}:
            out_path = Path("module_timeline.xlsx")
        else:
            out_path = Path(f"module_timeline_{suffix}.xlsx")

    return log_path, out_path


LOG_PATH, OUT_PATH = resolve_paths()

if not LOG_PATH.exists():
    raise FileNotFoundError(f"{LOG_PATH} not found")

pat = re.compile(r"^\[cycle\s+(\d+)\]\s+\[([^\]]+)\]\s+(.*)$")
transition_pat = re.compile(r"(.+?)\s*->\s*(.+?)(?:\s*\||$)")


def make_short_label(module, start_msg, end_msg):
    msg = start_msg

    # ------------------------
    # Scheduler-issued jobs
    # ------------------------
    if module == "SCHEDULER":
        if "issue NTT/FWD" in msg:
            m_src = re.search(r"src=([a-zA-Z0-9_]+)", msg)
            m_poly = re.search(r"poly=([0-9]+)", msg)
            src = m_src.group(1) if m_src else "?"
            poly = m_poly.group(1) if m_poly else ""

            if src == "z":
                return f"z{poly}"
            if src == "c":
                return "c"
            if src == "t1":
                return f"t1{poly}"
            return f"{src}{poly}"

        if "issue NTT/INTT" in msg:
            m_row = re.search(r"row=([0-9]+)", msg)
            row = m_row.group(1) if m_row else "?"
            return f"INTT{row}"

        if "issue PAU/mac_add_first" in msg:
            m_row = re.search(r"row=([0-9]+)", msg)
            row = m_row.group(1) if m_row else "?"
            return f"mac{row}"

        if "issue PAU/mac_add_acc" in msg:
            m_row = re.search(r"row=([0-9]+)", msg)
            m_col = re.search(r"col=([0-9]+)", msg)
            row = m_row.group(1) if m_row else "?"
            col = m_col.group(1) if m_col else "?"
            return f"acc{row},{col}"

        if "issue PAU/mac_sub" in msg:
            m_row = re.search(r"row=([0-9]+)", msg)
            row = m_row.group(1) if m_row else "?"
            return f"sub{row}"

        if "issue A generation" in msg:
            m_row = re.search(r"row=([0-9]+)", msg)
            m_col = re.search(r"col=([0-9]+)", msg)
            row = m_row.group(1) if m_row else "?"
            col = m_col.group(1) if m_col else "?"
            return f"A({row},{col})"

        if "issue POST/use_hint+pack" in msg:
            m_row = re.search(r"row=([0-9]+)", msg)
            row = m_row.group(1) if m_row else "?"
            return f"hint{row}"

        if "issue FINAL_HASH" in msg:
            return "c'"

        if "issue PREP/UNPACK" in msg:
            return "unpack"

        if "issue PREP/CHALLENGE_C" in msg:
            return "cgen"

        if "issue PREP/TR" in msg:
            return "tr"

        if "issue PREP/MU" in msg:
            return "mu"

    # ------------------------
    # Module-local logs
    # ------------------------
    if module == "NTTModule":
        if "start NTT" in msg:
            return "NTT"
        if "start INTT" in msg:
            return "INTT"
        if "PRE_MUL" in msg:
            return "pre"
        if "POST_MUL" in msg:
            return "post"

    if module == "PAU":
        if "mac_add_first" in msg:
            return "mac"
        if "mac_add_acc" in msg:
            return "acc"
        if "mac_sub" in msg:
            return "sub"

    if module == "HintPackModule":
        m_row = re.search(r"row=([0-9]+)", msg)
        row = m_row.group(1) if m_row else ""
        return f"hint{row}"

    if module == "SHAKE":
        if "ABSORB" in msg:
            if "tag=challenge_c" in msg:
                return "c-abs"
            if "tag=tr" in msg:
                return "tr-abs"
            if "tag=mu" in msg:
                return "mu-abs"
            if "tag=c_prime" in msg:
                return "c'-abs"
            m_row = re.search(r"row['\"]?:\s*([0-9]+)", msg)
            m_col = re.search(r"col['\"]?:\s*([0-9]+)", msg)
            if m_row and m_col:
                return f"Aa({m_row.group(1)},{m_col.group(1)})"
            return "ABS"

        if "PERMUTE" in msg:
            if "tag=challenge_c" in msg:
                return "c-perm"
            if "tag=tr" in msg:
                return "tr-perm"
            if "tag=mu" in msg:
                return "mu-perm"
            if "tag=c_prime" in msg:
                return "c'-perm"
            m_row = re.search(r"row['\"]?:\s*([0-9]+)", msg)
            m_col = re.search(r"col['\"]?:\s*([0-9]+)", msg)
            if m_row and m_col:
                return f"Ap({m_row.group(1)},{m_col.group(1)})"
            return "PERM"

        if "SQUEEZE" in msg:
            if "tag=challenge_c" in msg:
                return "c-sq"
            if "tag=tr" in msg:
                return "tr-sq"
            if "tag=mu" in msg:
                return "mu-sq"
            if "tag=c_prime" in msg:
                return "c'-sq"
            m_row = re.search(r"row['\"]?:\s*([0-9]+)", msg)
            m_col = re.search(r"col['\"]?:\s*([0-9]+)", msg)
            if m_row and m_col:
                return f"As({m_row.group(1)},{m_col.group(1)})"
            return "SQ"

        if "tag=challenge_c" in msg:
            return "cgen"
        if "tag=tr" in msg:
            return "tr"
        if "tag=mu" in msg:
            return "mu"
        if "tag=c_prime" in msg:
            return "c'"

        m_row = re.search(r"row['\"]?:\s*([0-9]+)", msg)
        m_col = re.search(r"col['\"]?:\s*([0-9]+)", msg)
        if m_row and m_col:
            return f"A({m_row.group(1)},{m_col.group(1)})"

        return "SHAKE"

    if module == "UniformSampler":
        m_row = re.search(r"row['\"]?:\s*([0-9]+)", msg)
        m_col = re.search(r"col['\"]?:\s*([0-9]+)", msg)
        if m_row and m_col:
            return f"Us({m_row.group(1)},{m_col.group(1)})"
        return "Usamp"

    if module == "SampleInBall":
        return "SIB"

    if module == "PkUnpacker":
        return "pk"
    if module == "SigUnpacker":
        return "sig"
    if module == "PackerModule":
        return "pack"

    return module[:6]


events = []
with LOG_PATH.open("r", encoding="utf-8") as f:
    for line in f:
        line = line.rstrip("\n")
        m = pat.match(line)
        if m:
            cycle = int(m.group(1))
            module = m.group(2).strip()
            msg = m.group(3).strip()
            events.append((cycle, module, msg))

# Stable module order from Events tab (first appearance)
event_modules = list(dict.fromkeys(module for _, module, _ in events))
colors = [
    "5B9BD5", "ED7D31", "70AD47", "FFC000", "4472C4",
    "A5A5A5", "C00000", "00B0F0", "92D050", "7030A0"
]
module_fills = {
    m: PatternFill("solid", fgColor=colors[i % len(colors)])
    for i, m in enumerate(event_modules)
}

# Build intervals from state transitions
intervals = []
active = {}

for cycle, module, msg in events:
    tm = transition_pat.match(msg)
    if not tm:
        continue

    old_state = tm.group(1).strip()
    new_state = tm.group(2).strip()

    start_states = {"BUSY", "RUN", "ABSORB", "PRE_MUL", "CORE", "PERMUTE", "SQUEEZE", "POST_MUL"}
    if old_state == "IDLE" and new_state in start_states:
        active.setdefault(module, []).append((cycle, msg))
    elif new_state == "IDLE":
        if active.get(module):
            start_cycle, start_msg = active[module].pop(0)
            duration = max(1, cycle - start_cycle)
            label = make_short_label(module, start_msg, msg)
            intervals.append([module, start_cycle, duration, start_msg, msg, label])

last_cycle = max([e[0] for e in events], default=0)

wb = Workbook()

# ----------------------------------------------------------------------
# Events sheet
# ----------------------------------------------------------------------
ws_events = wb.active
ws_events.title = "Events"
ws_events.append(["Cycle", "Module", "Message"])
for row in events:
    ws_events.append(list(row))

# Color event rows by module
for r in range(2, ws_events.max_row + 1):
    module = ws_events.cell(row=r, column=2).value
    fill = module_fills.get(module)
    if fill:
        ws_events.cell(row=r, column=2).fill = fill

for cell in ws_events[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center")

ws_events.column_dimensions["A"].width = 12
ws_events.column_dimensions["B"].width = 18
ws_events.column_dimensions["C"].width = 100
ws_events.freeze_panes = "A2"

# ----------------------------------------------------------------------
# Intervals sheet
# ----------------------------------------------------------------------
ws_int = wb.create_sheet("Intervals")
ws_int.append(["Module", "Start Cycle", "Duration", "Start Message", "End Message", "Label"])
for row in intervals:
    ws_int.append(row)

for cell in ws_int[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="E2F0D9")
    cell.alignment = Alignment(horizontal="center")

ws_int.column_dimensions["A"].width = 18
ws_int.column_dimensions["B"].width = 14
ws_int.column_dimensions["C"].width = 12
ws_int.column_dimensions["D"].width = 80
ws_int.column_dimensions["E"].width = 80
ws_int.column_dimensions["F"].width = 16
ws_int.freeze_panes = "A2"

# ----------------------------------------------------------------------
# Summary + Gantt chart
# ----------------------------------------------------------------------
ws_sum = wb.create_sheet("Summary")
ws_sum["A1"] = "Module timeline summary"
ws_sum["A1"].font = Font(bold=True, size=14)
ws_sum["A3"] = "Total logged events"
ws_sum["B3"] = len(events)
ws_sum["A4"] = "Total intervals"
ws_sum["B4"] = len(intervals)
ws_sum["A5"] = "Last logged cycle"
ws_sum["B5"] = last_cycle
ws_sum["A7"] = "Note"
ws_sum["B7"] = f"Chart uses module state transitions from {LOG_PATH}."

ws_g = wb.create_sheet("GanttData")
ws_g.append(["Task", "Start", "Duration", "Module", "Label"])

mod_counts = {}
for module, start, duration, s_msg, e_msg, label in intervals:
    mod_counts[module] = mod_counts.get(module, 0) + 1
    task_name = f"{module} #{mod_counts[module]} ({label})"
    ws_g.append([task_name, start, duration, module, label])

for cell in ws_g[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="FCE4D6")

if intervals:
    chart = BarChart()
    chart.type = "bar"
    chart.style = 2
    chart.title = "Module timeline (Gantt-like)"
    chart.y_axis.title = "Tasks"
    chart.x_axis.title = "Cycle"
    chart.height = 20
    chart.width = 34

    start_ref = Reference(ws_g, min_col=2, min_row=1, max_row=1 + len(intervals))
    dur_ref = Reference(ws_g, min_col=3, min_row=1, max_row=1 + len(intervals))
    cats = Reference(ws_g, min_col=1, min_row=2, max_row=1 + len(intervals))

    start_series = Series(start_ref, title_from_data=True)
    dur_series = Series(dur_ref, title_from_data=True)

    chart.series.append(start_series)
    chart.series.append(dur_series)
    chart.set_categories(cats)

    # Hide start offset series
    start_series.graphicalProperties.noFill = True
    start_series.graphicalProperties.line.noFill = True

    ws_sum.add_chart(chart, "A10")

# ----------------------------------------------------------------------
# Timeline grid sheet (coarse bins)
# ----------------------------------------------------------------------
ws_grid = wb.create_sheet("TimelineGrid")
bin_size = 200
num_bins = max(1, math.ceil(last_cycle / bin_size))

ws_grid["A1"] = f"Timeline grid ({bin_size}-cycle bins)"
ws_grid["A1"].font = Font(bold=True)

ws_grid["A2"] = "Module"
for i in range(num_bins + 1):
    ws_grid.cell(row=2, column=i + 2, value=i * bin_size)

for cell in ws_grid[2]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center")

modules = sorted(set(module for module, *_ in intervals))
modules = event_modules

for r, module in enumerate(modules, start=3):
    ws_grid.cell(row=r, column=1, value=module)
    ws_grid.cell(row=r, column=1).fill = module_fills[module]

    for m, start, duration, _, _, label in intervals:
        if m != module:
            continue

        end = start + duration
        start_bin = start // bin_size
        end_bin = end // bin_size

        for b in range(start_bin, end_bin + 1):
            c = ws_grid.cell(row=r, column=b + 2)
            c.fill = module_fills[module]

            if c.value is None:
                c.value = label
            else:
                existing = str(c.value).split("/")
                if label not in existing:
                    c.value = str(c.value) + "/" + label

for col in range(1, num_bins + 3):
    ws_grid.column_dimensions[get_column_letter(col)].width = 14 if col > 1 else 18

for row in ws_grid.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

ws_grid.freeze_panes = "B3"

wb.save(OUT_PATH)
print(f"Created {OUT_PATH}")
